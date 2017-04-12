#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
from openpyxl import load_workbook
from openpyxl import Workbook

try:
  file_name = str(sys.argv[1])
except IndexError:
  exit("Usage: " + sys.argv[0] + " filename.xlsx")

workbook = load_workbook(file_name)

sheet_names = workbook.get_sheet_names()

for sheet in sheet_names:
  current_sheet = workbook.get_sheet_by_name(sheet)
  sheet_comments = []

  # Names the sheet in the output using a value at a cell (A3 in this case)
  # Change this for how you want to refer to the each sheet.
  # Another possibility is to pick a cell that appears on each sheet and has identifying
  # info for that sheet example: current_sheet['A1'].value
  sheet_label_reference = current_sheet.title

  for row in current_sheet.iter_rows():
    for cell in row:
      comment = cell.comment
      if comment:
        comment_lines = comment.text.split('\n')[1:]
        print (comment_lines)
        for line in comment_lines:
          sheet_comments.append(line)

if sheet_comments:
  print(sheet_label_reference,":", " ".join(str(x) for x in sheet_comments))