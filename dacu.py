#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
from openpyxl import load_workbook
from openpyxl import Workbook

def process(sourcefile, destfile):
  #open source excel file and get data
  sourcebook = load_workbook(sourcefile)
  sourcesheets = sourcebook.get_sheet_names()
  sourcesheet = sourcebook.get_sheet_by_name(sourcesheets[0])
  rows = sourcesheet.rows
  columns = sourcesheet.columns
  i = 0
  j = 0
  for row in rows:
    i += 1
  for column in columns:
    j += 1

  #create dest excel, prepare to write data into it
  destbook = Workbook()
  destsheet = destbook.get_active_sheet()
  print (destsheet.title)
  destsheet.title = 'NiuXinxin'
  print(destsheet.title)
  for row in range(1, i):
    comments = ""
    for column in range(1, j):
      destsheet.cell(row=row, column=column).value = sourcesheet.cell(row=row, column=column).value
      if sourcesheet.cell(row=row, column=column).comment:
        print(sourcesheet.cell(row=row, column=column).comment)
        comments += str(sourcesheet.cell(row=row, column=column).comment)
    if comments:
      destsheet.cell(row=row, column=j).value = comments
  destbook.save(filename='new_file.xlsx')


if __name__ == "__main__":
  sourcefile = str(sys.argv[1])
  destfile = str(sys.argv[2])
  print (sourcefile)
  print (destfile)
  process(sourcefile, destfile)